"""
Update the Joins sheet in the Excel file to match the new knowledge graph structure
"""

import pandas as pd
from openpyxl import load_workbook

# Read the existing Excel file
excel_file = "AI_SampleDataStruture.xlsx"

# Create the updated Joins data
joins_data = {
    "Table1": ["Counterparty", "Counterparty", "Counterparty", "Counterparty"],
    "Table2": ["Trade", "Concentration", "Concentration", "Concentration"],
    "Join Key Table1": [
        "Entity+Counterparty ID",
        "Entity+Counterparty Country",
        "Entity+Counterparty Sector",
        "Entity+Internal Rating"
    ],
    "Join Key Table2": [
        "Entity+Reporting Counterparty ID",
        "Entity+Concentration Value",
        "Entity+Concentration Value",
        "Entity+Concentration Value"
    ],
    "__EMPTY": [
        "default",
        "For country level data",
        "For sector level data",
        "For rating level data"
    ]
}

# Create DataFrame
joins_df = pd.DataFrame(joins_data)

# Load the workbook
wb = load_workbook(excel_file)

# Remove the old Joins sheet if it exists
if "Joins" in wb.sheetnames:
    del wb["Joins"]

# Create new Joins sheet
ws = wb.create_sheet("Joins")

# Write headers
for col_idx, column_name in enumerate(joins_df.columns, start=1):
    ws.cell(row=1, column=col_idx, value=column_name)

# Write data
for row_idx, row in enumerate(joins_df.values, start=2):
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Save the workbook
wb.save(excel_file)

print("✅ Updated Joins sheet in Excel file!")
print("\nNew relationships:")
for idx, row in joins_df.iterrows():
    print(f"{idx + 1}. {row['Table1']} → {row['Table2']}")
    print(f"   Keys: {row['Join Key Table1']} = {row['Join Key Table2']}")
    print(f"   Context: {row['__EMPTY']}")
    print()
